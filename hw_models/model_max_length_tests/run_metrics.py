#########################################################################################
# Copyright (c) Qualcomm Technologies, Inc. and/or its subsidiaries. All rights reserved.
# Confidential and Proprietary - Qualcomm Technologies, Inc. and/or its subsidiaries.
#########################################################################################

"""Per-case timings, the optional prefill regression check, and the run summary table.

REF_PERF follows the convention of the other suites in this directory: an absent key
prints a notice rather than failing, so references can be populated from a first clean
run.

A case may be a batch. Timings are then reported as the batch sees them: prefill is the
slowest request's (the point at which the whole batch has its first token) and decode
throughput is summed across requests, which is the number that improves with batching
even though per-request throughput falls.
"""

import contextlib
import fcntl
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from tabulate import tabulate

from ctx_config import PERF_TOLERANCE
from prompt_builder import BuiltInput

REF_PERF: dict[tuple, dict] = {
    # (model_name, tp_size, ctx_len, modality, batch_size): {"prefill": seconds}
}

# Written once above the table, since the columns need more explanation than their
# headers have room for.
SUMMARY_LEGEND = (
    "BS is the number of requests issued together; with BS>1 the requests are\n"
    "distinct prompts of the same length, Prompt Tok/Decode Tok and the token\n"
    "breakdown describe one of them, Prefill is the slowest request's and Decode\n"
    "tok/s is summed across the batch. Prompt Tok is what the engine actually saw;\n"
    "Decode Tok is how many tokens that request generated. Text/Visual/Img/Vid Tok\n"
    "are the build-time breakdown: Text Tok (scaffold + filler + question) is\n"
    "exact, and Text + Visual is the predicted prompt length. Images and Video give\n"
    "the synthetic inputs as count x WxH and count x framesf WxH respectively (one\n"
    "clip's frame count, not the sum, since every clip in a case is the same size).\n"
    "A video processor may apply its own pixel budget, so a video row's Prompt Tok\n"
    "can drift from Text + Visual."
)

_summary_rows: list[dict] = []


def _describe_images(built: BuiltInput) -> str:
    if not built.num_images or built.image_size is None:
        return "-"
    width, height = built.image_size
    return f"{built.num_images}x {width}x{height}"


def _describe_video(built: BuiltInput) -> str:
    if not built.num_videos or built.video_size is None:
        return "-"
    width, height = built.video_size
    return f"{built.num_videos}x {built.num_frames}f {width}x{height}"


def _timings(results, decode_token_counts: list[int], wall_time: float):
    """(prefill, decode_tps) for the batch, or a wall-clock fallback.

    prefill is the max across requests because that is when the batch as a whole has
    produced its first tokens; decode throughput is the sum because that is the
    quantity batching is supposed to increase.
    """
    prefills, decode_rates = [], []
    for index, (result, decode_tokens) in enumerate(
        zip(results, decode_token_counts, strict=True)
    ):
        metrics = getattr(result, "metrics", None)
        if metrics is None:
            return None, 0.0
        print(
            f"    [{index}] scheduled_ts={metrics.scheduled_ts:.3f} "
            f"first_token_ts={metrics.first_token_ts:.3f} "
            f"last_token_ts={metrics.last_token_ts:.3f} "
            f"prefill: {(metrics.first_token_ts - metrics.scheduled_ts):.3f}"
        )
        # first_token_latency is the end-user TTFT (arrival -> first token);
        # prefill excludes host-side preprocessing and queueing, so their
        # difference recovers the host-side portion.
        prefills.append(metrics.first_token_ts - metrics.scheduled_ts)
        decode_time = metrics.last_token_ts - metrics.first_token_ts
        decode_rates.append(
            (decode_tokens - 1) / decode_time if decode_time > 0 else 0.0
        )

    slowest = max(prefills)
    ttft = max(getattr(r, "metrics").first_token_latency for r in results)
    print(
        f"TTFT={ttft:.2f}s prefill={slowest:.2f}s "
        f"rendering={ttft - slowest:.2f}s "
        f"decode_tps={sum(decode_rates):.2f} total"
        + (
            f" ({sum(decode_rates) / len(decode_rates):.2f}/request)"
            if len(decode_rates) > 1
            else ""
        )
    )
    return slowest, sum(decode_rates)


def record_metrics(
    model_name: str,
    tp_size: int,
    modality_name: str,
    ctx_len: int,
    engine_len: int,
    results,
    wall_time: float,
    per_request: list[tuple[int, int]],
    batch: list[BuiltInput],
) -> float | None:
    """Print per-case timings and stash a summary row. Returns prefill seconds."""
    decode_token_counts = [decode for _prompt, decode in per_request]
    prefill, decode_tps = _timings(results, decode_token_counts, wall_time)

    if prefill is None:
        print(
            "RequestOutput.metrics is None -- expected disable_log_stats=False on "
            "the LLM to populate it. Falling back to wall clock."
        )
        prefill = wall_time

    built = batch[0]
    _summary_rows.append(
        {
            "Model": model_name.split("/")[-1],
            "TP": tp_size,
            "BS": len(batch),
            "Modality": modality_name,
            "Ctx": ctx_len,
            "Engine Len": engine_len,
            "Prompt Tok": per_request[0][0],
            "Decode Tok": per_request[0][1],
            "Text Tok": built.text_tokens,
            "Filler Tok": built.filler_tokens,
            "Visual Tok": built.visual_tokens,
            "Img Tok": built.image_tokens,
            "Vid Tok": built.video_tokens,
            "Images": _describe_images(built),
            "Video": _describe_video(built),
            "Prefill (s)": prefill or 0.0,
            "Decode tok/s": decode_tps,
            "Wall (s)": wall_time,
        }
    )
    return prefill


def check_perf_reference(
    model_name: str,
    tp_size: int,
    ctx_len: int,
    modality_name: str,
    prefill: float | None,
    batch_size: int = 1,
) -> None:
    key = (model_name, tp_size, ctx_len, modality_name, batch_size)
    reference = REF_PERF.get(key)
    if reference is None:
        print(f"Missing performance reference for {key}")
        return
    if prefill is None:
        return
    ref_prefill = reference["prefill"]
    delta_pct = (prefill - ref_prefill) / ref_prefill
    # A negative delta is a speed-up; only regressions fail.
    assert delta_pct <= PERF_TOLERANCE, (
        f"Prefill regression for {key}: measured={prefill:.2f}s "
        f"reference={ref_prefill:.2f}s delta={delta_pct:+.2%} "
        f"allowed={PERF_TOLERANCE:.2%}"
    )
    print(f"Prefill within tolerance for {key}: {prefill:.2f}s ({delta_pct:+.2%})")


def render_summary() -> str | None:
    """The collected rows as a grid table, or None if nothing ran."""
    if not _summary_rows:
        return None
    headers = list(_summary_rows[0].keys())
    table = [
        [
            f"{value:.2f}" if isinstance(value, float) else value
            for value in row.values()
        ]
        for row in _summary_rows
    ]
    return tabulate(table, headers=headers, tablefmt="grid")


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)


@contextlib.contextmanager
def _locked_for_append(path: Path):
    """Cross-process lock on a sidecar `<path>.lock` file.

    A single sequential run never contends for this, but scheduler.py can have
    several pytest processes alive at once, each wanting to append its own rows to
    the same run's workbook -- without this, two concurrent load-append-save cycles
    race and the loser's rows are silently dropped.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = path.with_name(path.name + ".lock")
    with open(lock_path, "w") as lock_file:
        fcntl.flock(lock_file, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lock_file, fcntl.LOCK_UN)


def append_summary_workbook(path: Path) -> None:
    """Append this session's rows to an .xlsx, creating it (with a Legend sheet) if new.

    Mirrors the append semantics the suite used for its old plain-text summary: one
    workbook per run directory, one "Summary" sheet that every pytest process in the
    run appends its rows to. Concurrent writers (scheduler.py) serialize on
    _locked_for_append rather than racing on the read-modify-write below.
    """
    if not _summary_rows:
        return

    with _locked_for_append(path):
        if path.exists() and path.stat().st_size > 0:
            workbook = load_workbook(path)
            sheet = workbook["Summary"]
            headers = [cell.value for cell in sheet[1]]
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            headers = list(_summary_rows[0].keys())
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Summary"
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            sheet.freeze_panes = "A2"

            legend = workbook.create_sheet("Legend")
            for line in SUMMARY_LEGEND.splitlines():
                legend.append([line])
            legend.column_dimensions["A"].width = 100

        for row in _summary_rows:
            sheet.append([row.get(header, "") for header in headers])

        _autosize(sheet)
        workbook.save(path)
